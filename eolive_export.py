'''
    #Version : 0.1
    #Last modified : 17-04-2018
    #Modified By : Hari
    #Purpose : This script is useful in extracting the configuration details of all the eolive charts if eolive export is available
    #prerequisites : Python 2 or 3. Idle tool that comes with python installation and openpyxl
    #Compatibility : This script is so far compatible only for specific customers and subject to consistent changes as chart configuration varies greatly among customers
    #Target Environment : Windows
    #Python version used : 3.6.5
'''

from openpyxl.styles import PatternFill
import json,os,openpyxl
import re

def main():
    excel = openpyxl.Workbook() #Open the empty excel workbook
    active_sheet = excel.active #Get the current active worksheet
    name = active_sheet.title
    sheet_name = excel[name]
    sheet_name.title = "Dataviews_Charts" #Name the worksheet as per your need

    #Title names of each columns in our output excel sheet. Change it as you wish
    active_sheet['A1'] = "Dataviews"
    active_sheet['B1'] = "Charts"
    active_sheet['C1'] = "Input Filters"
    active_sheet['D1'] = "KPI Expression"
    active_sheet['E1'] = "GroupBy Values"
    active_sheet['F1'] = "Granularity"
    active_sheet['G1'] = "Chart Type"
    active_sheet['H1'] = "Value Axis"
    active_sheet['I1'] = "Value Scale"
    active_sheet['J1'] = "Ranking"
    active_sheet['K1'] = "drill Options"

    #Using Yellow colour to highlight the headings
    YellowFill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    active_sheet['A1'].fill = YellowFill
    active_sheet['B1'].fill = YellowFill
    active_sheet['C1'].fill = YellowFill
    active_sheet['D1'].fill = YellowFill
    active_sheet['E1'].fill = YellowFill
    active_sheet['F1'].fill = YellowFill
    active_sheet['G1'].fill = YellowFill
    active_sheet['H1'].fill = YellowFill
    active_sheet['I1'].fill = YellowFill
    active_sheet['J1'].fill = YellowFill
    active_sheet['K1'].fill = YellowFill   
    

    #Initializing the variables which will be used later
    j = 2
    filter3 = " "
    ranking = " "
    group_by = " "
    value_axis = " "
    chart_type = " "
    granularity = " " 
    scale_type = " "
    group_by_val = " "
    drill_kpi = [] #There might be more than one drill configured, so we are using the List data structure to append all the values


    #rootdir is the location of the eolive export. Don't forget to give the path until the public folder.
    rootdir = 'path of the source folder'

    #os.walk is used to walk the rootdir in top-down approach. it visits all the directories
    for currDir, children, files in os.walk(rootdir):
        for f in files:
            if not f.startswith('.'):   #We are eliminating files starting with '.'
                i = 0
                my_file = currDir+'\\'+f #This is the full file path
                f_ptr = open(my_file)
                lines = f_ptr.readlines()
                for line in lines:
                    dataview = currDir.split('\\')[-1]
                    ex_DV = "A" + str(j)
                    ex_CH = "B" + str(j)
                    ex_Inp_FILTER = "C" + str(j)
                    ex_Kpi_Exp = "D" + str(j)
                    ex_group = "E" + str(j)
                    ex_granular = "F" + str(j)
                    ex_Chart_Type = "G" + str(j)
                    ex_Val_Axis = "H" + str(j)
                    ex_Val_Scale = "I" + str(j)
                    ex_Ranking = "J" + str(j)
                    ex_drill = "K" + str(j)
                    i = i+1

                    #Below lines are to check whether drill configured or not
                    if 'drill_kpis {' in line:
                        if (('OperatorE212' in lines[i]) or ('OperatorPrefix' in lines[i])) and (('mcls_lookup_prefix' in lines[i])|('mcls_lookup' in lines[i])):
                            if 'Based on Operator Name or Country' not in drill_kpi:
                                drill_kpi.append("Based on Operator Name or Country")

                        elif ('IMSIPrefix' in lines[i]) and (('mcls_lookup_prefix' in lines[i])|('mcls_lookup' in lines[i])):
                            if 'Subscriber Prefix Name/IMSI Prefix' not in drill_kpi:
                                drill_kpi.append('Subscriber Prefix Name/IMSI Prefix')
                        else:
                            drill_kpi.append(lines[i].replace("field_name:",""))

                    #To find the TDR format used for the chart
                    if 'format_name:' in line:
                        my_format = lines[i-1]
                        final_format = my_format.replace("format_name:"," ")

                    #To find the input filter configured
                    if 'input_filter {' in line:
                        my_filter = lines[i]
                        filter1 = my_filter.replace("\\n", " ")
                        filter2 = filter1.replace("\\"," ")
                        filter3 = filter2.replace("user:","")
                        
                    #To find the KPI_Expression configured
                    if 'kpi_expression {' in line:
                        kpi_exp = lines[i]
                        kpi_filter1 = kpi_exp.replace("\\n", " ")
                        kpi_filter2 = kpi_filter1.replace("\\"," ")
                        kpi_filter3 = kpi_filter2.replace("user:","")

                    #To find the "Group By" values configured
                    if 'main_kpi {' in line:
                        if "GRXLinksets" in lines[i]:
                            group_by_val = "GRXCarrier"
                        elif (('OperatorE212' in lines[i]) or ('OperatorPrefix' in lines[i])) and (('mcls_lookup_prefix' in lines[i])|('mcls_lookup' in lines[i])):
                            group_by_val = 'Based on Operator Name or Country'
                        else:
                            group_by_val = lines[i].split(':')[-1]

                    #The below lines are to find Ranking, Chart Type, Scale Type                         
                    if 'main_kpi {' in line:
                        ranking = lines[i+4].split(':')[-1] + lines[i+5].split(':')[-1]
                     
                    if 'value_axis_label:' in line:
                        value_axis = lines[i-1].split(':')[-1]
                        
                    if 'chart_properties {' in line:
                        chart_type = lines[i+2].split(':')[-1]
                        
                    if 'granularity:' in line:
                        granularity = lines[i-1].split(':')[-1]
                        
                    if 'scaleType:' in line:
                        scale_type = lines[i-1].split(':')[-1]

                                    
                #The below lines are for the filling of fields in excel sheet generated
                active_sheet[ex_DV] = dataview
                active_sheet[ex_CH] = f
                active_sheet[ex_Inp_FILTER] = filter3
                active_sheet[ex_Kpi_Exp] = kpi_filter3
                active_sheet[ex_group] = group_by_val
                active_sheet[ex_granular] = granularity
                active_sheet[ex_Chart_Type] = chart_type
                active_sheet[ex_Val_Axis] = value_axis
                active_sheet[ex_Val_Scale] = scale_type
                if 'TOP' in ranking or 'Bottom' in ranking:
                    active_sheet[ex_Ranking] = ranking
                else:
                    active_sheet[ex_Ranking] = "No Ranking Enabled"
                if not drill_kpi:
                    active_sheet[ex_drill] = "No Drill Configured"
                else:
                    active_sheet[ex_drill] = ' '.join(drill_kpi)

                filter3 = " "
                ranking = " "
                group_by = " "
                value_axis = " "
                chart_type = " "
                granularity = " " 
                scale_type = " "
                group_by_val = " "
                kpi_filter3 = " "
                filter3 = " "
                drill_kpi = []
                j = j+1                                            

    #Save your excel sheet via below line. Don't keep the excel open while executing the script
    excel.save("path of destination folder\file_name.xlsx")

if __name__ == "__main__":
    main()
